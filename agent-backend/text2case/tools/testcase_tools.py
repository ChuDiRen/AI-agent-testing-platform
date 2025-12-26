"""测试用例生成工具模块

提供测试用例生成、验证、导出等工具函数
这些工具可以被 ReAct Agent 调用，实现工具驱动的测试用例生成
"""
import json
import re
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from difflib import SequenceMatcher
from enum import Enum

from langchain_core.tools import tool


# ============== 测试方法选择工具 ==============

class TestMethodType(Enum):
    """测试方法类型"""
    EQUIVALENCE_CLASS = "equivalence_class"
    BOUNDARY_VALUE = "boundary_value"
    DECISION_TABLE = "decision_table"
    SCENARIO = "scenario"
    ORTHOGONAL = "orthogonal"
    CAUSE_EFFECT = "cause_effect"


METHOD_TEMPLATES = {
    TestMethodType.EQUIVALENCE_CLASS: """
## 等价类划分法
- 识别所有输入参数
- 为每个参数划分有效等价类和无效等价类
- 从每个等价类选取代表值
- 优先级: 有效等价类P0, 无效等价类P1
""",
    TestMethodType.BOUNDARY_VALUE: """
## 边界值分析法
- 识别所有有边界的输入参数
- 确定边界点: min, max
- 设计边界值: min-1, min, min+1, max-1, max, max+1
- 边界点必须覆盖: P0
""",
    TestMethodType.DECISION_TABLE: """
## 判定表法
- 识别所有条件（输入）和动作（输出）
- 列出所有条件组合
- 确定每种组合对应的动作
- 每条规则对应一个用例
""",
    TestMethodType.SCENARIO: """
## 场景法
- 识别基本流（Happy Path）
- 识别备选流（正常分支）
- 识别异常流（错误处理）
- 基本流必须完整覆盖: P0
""",
    TestMethodType.ORTHOGONAL: """
## 正交法
- 识别所有因素（参数）和水平（取值）
- 选择合适的正交表
- 按正交表组合设计用例
- 用N个用例覆盖N^k种组合
""",
    TestMethodType.CAUSE_EFFECT: """
## 因果图法
- 识别原因（输入条件）和结果（输出动作）
- 分析因果关系和约束
- 转换为判定表
- 覆盖所有因果组合
""",
}

FEATURE_KEYWORDS = {
    TestMethodType.EQUIVALENCE_CLASS: {
        "输入": 2, "验证": 2, "校验": 2, "格式": 2, "类型": 1,
        "有效": 2, "无效": 2, "用户名": 1, "密码": 1, "邮箱": 1,
    },
    TestMethodType.BOUNDARY_VALUE: {
        "范围": 3, "长度": 3, "大小": 2, "最大": 3, "最小": 3,
        "边界": 3, "限制": 2, "字符": 2, "数量": 2, "金额": 2,
    },
    TestMethodType.DECISION_TABLE: {
        "条件": 3, "规则": 3, "逻辑": 2, "判断": 2, "如果": 2,
        "否则": 2, "并且": 2, "或者": 2, "权限": 2, "状态": 1,
    },
    TestMethodType.SCENARIO: {
        "流程": 3, "步骤": 2, "场景": 3, "操作": 1, "业务": 2,
        "购买": 2, "下单": 2, "注册": 2, "登录": 2, "接口": 1,
    },
    TestMethodType.ORTHOGONAL: {
        "组合": 3, "配置": 3, "参数": 1, "选项": 2, "设置": 2,
        "多个": 2, "搭配": 2, "兼容": 2,
    },
    TestMethodType.CAUSE_EFFECT: {
        "依赖": 3, "关联": 3, "互斥": 3, "约束": 3, "前提": 2,
        "制约": 2, "影响": 2, "触发": 2,
    },
}


@tool
def select_test_methods(requirement: str, max_methods: int = 2) -> Dict[str, Any]:
    """根据需求自动选择最适合的测试方法
    
    分析需求特征，返回推荐的测试方法和对应的模板指导。
    这是一个确定性工具，不消耗 LLM Token。
    
    Args:
        requirement: 需求描述文本
        max_methods: 最多选择几种方法，默认2种
        
    Returns:
        包含推荐方法、模板和匹配分数的字典
    """
    scores = {}
    req_lower = requirement.lower()
    
    for method_type, keywords in FEATURE_KEYWORDS.items():
        score = sum(weight for kw, weight in keywords.items() if kw in req_lower)
        scores[method_type] = score
    
    # 按分数排序选择
    sorted_methods = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    selected = [m for m, s in sorted_methods if s >= 2][:max_methods]
    
    if not selected:
        selected = [TestMethodType.SCENARIO]
    
    # 构建模板
    templates = []
    method_names = {
        TestMethodType.EQUIVALENCE_CLASS: "等价类划分法",
        TestMethodType.BOUNDARY_VALUE: "边界值分析法",
        TestMethodType.DECISION_TABLE: "判定表法",
        TestMethodType.SCENARIO: "场景法",
        TestMethodType.ORTHOGONAL: "正交法",
        TestMethodType.CAUSE_EFFECT: "因果图法",
    }
    
    for method in selected:
        templates.append(METHOD_TEMPLATES.get(method, ""))
    
    return {
        "recommended_methods": [method_names[m] for m in selected],
        "method_ids": [m.value for m in selected],
        "templates": "\n---\n".join(templates),
        "scores": {method_names[m]: scores[m] for m in selected},
    }


# ============== 测试数据生成工具 ==============

@tool
def generate_test_data(
    field_name: str,
    field_type: str,
    constraints: Optional[str] = None
) -> Dict[str, Any]:
    """生成测试数据
    
    根据字段类型和约束条件，自动生成有效和无效的测试数据。
    
    Args:
        field_name: 字段名称（如 username, email, age）
        field_type: 字段类型（string, integer, email, phone, date）
        constraints: 约束条件（如 "min:1, max:100, required"）
        
    Returns:
        包含有效数据、无效数据、边界数据的字典
    """
    result = {
        "field": field_name,
        "type": field_type,
        "valid_data": [],
        "invalid_data": [],
        "boundary_data": [],
    }
    
    # 解析约束
    min_val, max_val, required = None, None, False
    if constraints:
        if "min:" in constraints:
            min_val = int(re.search(r'min:(\d+)', constraints).group(1))
        if "max:" in constraints:
            max_val = int(re.search(r'max:(\d+)', constraints).group(1))
        required = "required" in constraints
    
    if field_type == "string":
        result["valid_data"] = ["test_value", "中文测试", "Test123"]
        result["invalid_data"] = ["", " ", None] if required else [None]
        if max_val:
            result["boundary_data"] = [
                "a" * (max_val - 1),
                "a" * max_val,
                "a" * (max_val + 1),
            ]
            result["invalid_data"].append("a" * (max_val + 100))
    
    elif field_type == "integer":
        min_val = min_val or 0
        max_val = max_val or 100
        result["valid_data"] = [min_val + 1, (min_val + max_val) // 2, max_val - 1]
        result["invalid_data"] = ["abc", None, 1.5, ""]
        result["boundary_data"] = [
            min_val - 1, min_val, min_val + 1,
            max_val - 1, max_val, max_val + 1,
        ]
    
    elif field_type == "email":
        result["valid_data"] = ["test@example.com", "user.name@domain.cn"]
        result["invalid_data"] = ["invalid", "@example.com", "test@", "test@.com", ""]
        result["boundary_data"] = ["a@b.cn", "a" * 50 + "@example.com"]
    
    elif field_type == "phone":
        result["valid_data"] = ["13800138000", "18612345678"]
        result["invalid_data"] = ["1380013800", "138001380001", "12345678901", "abcdefghijk"]
        result["boundary_data"] = ["10000000000", "19999999999"]
    
    elif field_type == "date":
        result["valid_data"] = ["2024-01-15", "2024-12-31"]
        result["invalid_data"] = ["2024-13-01", "2024-02-30", "invalid", ""]
        result["boundary_data"] = ["1970-01-01", "2099-12-31"]
    
    return result


# ============== 测试用例验证工具 ==============

@tool
def validate_testcase_format(testcase_text: str) -> Dict[str, Any]:
    """验证测试用例格式是否规范
    
    检查测试用例是否包含必要字段，格式是否正确。
    
    Args:
        testcase_text: 测试用例文本（Markdown格式）
        
    Returns:
        验证结果，包含是否有效、问题列表、统计信息
    """
    issues = []
    stats = {
        "total_cases": 0,
        "p0_count": 0,
        "p1_count": 0,
        "p2_count": 0,
        "p3_count": 0,
    }
    
    # 查找所有用例
    case_pattern = r'#{2,4}\s*(TC-?\d+)\s*([^\n]*)'
    cases = re.findall(case_pattern, testcase_text)
    stats["total_cases"] = len(cases)
    
    if not cases:
        issues.append("未找到测试用例（格式应为: ### TC-001 用例标题）")
        return {"valid": False, "issues": issues, "stats": stats}
    
    # 检查必要字段
    required_fields = ["优先级", "测试步骤", "预期结果"]
    for field in required_fields:
        if field not in testcase_text:
            issues.append(f"缺少必要字段: {field}")
    
    # 统计优先级分布
    p0_matches = re.findall(r'\*\*优先级\*\*[：:]\s*P0', testcase_text)
    p1_matches = re.findall(r'\*\*优先级\*\*[：:]\s*P1', testcase_text)
    p2_matches = re.findall(r'\*\*优先级\*\*[：:]\s*P2', testcase_text)
    p3_matches = re.findall(r'\*\*优先级\*\*[：:]\s*P3', testcase_text)
    
    stats["p0_count"] = len(p0_matches)
    stats["p1_count"] = len(p1_matches)
    stats["p2_count"] = len(p2_matches)
    stats["p3_count"] = len(p3_matches)
    
    # 检查优先级分布是否合理
    if stats["p0_count"] == 0:
        issues.append("警告: 没有P0级别的用例，核心功能可能未覆盖")
    
    if stats["p0_count"] > stats["total_cases"] * 0.5:
        issues.append("警告: P0用例过多（>50%），优先级可能划分不合理")
    
    return {
        "valid": len([i for i in issues if not i.startswith("警告")]) == 0,
        "issues": issues,
        "stats": stats,
    }


# ============== 导出工具 ==============

@tool
def export_to_xmind(testcase_text: str, output_dir: Optional[str] = None) -> Dict[str, Any]:
    """将测试用例导出为XMind思维导图
    
    解析测试用例文本，生成XMind格式的思维导图文件。
    
    Args:
        testcase_text: 测试用例文本（Markdown格式）
        output_dir: 输出目录，默认为 text2case/output
        
    Returns:
        导出结果，包含文件路径和统计信息
    """
    from ..models import TestCaseSuite, TestCaseModule, TestCaseItem
    
    # 解析测试用例
    cases_list = []
    case_pattern = r'#{2,4}\s*(TC-?\d+)\s*([^\n]*)'
    headers = list(re.finditer(case_pattern, testcase_text))
    
    for idx, match in enumerate(headers):
        case_id = match.group(1)
        case_title = match.group(2).strip() or case_id
        
        start = match.end()
        end = headers[idx + 1].start() if idx + 1 < len(headers) else len(testcase_text)
        content = testcase_text[start:end]
        
        # 提取字段
        priority_match = re.search(r'\*\*优先级\*\*[：:]\s*(\S+)', content)
        priority = priority_match.group(1) if priority_match else "P1"
        
        steps = re.findall(r'\d+\.\s*([^\n]+)', 
            re.search(r'\*\*测试步骤\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL).group(1)
            if re.search(r'\*\*测试步骤\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL) else "")
        
        expected = re.findall(r'\d+\.\s*([^\n]+)',
            re.search(r'\*\*预期结果\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL).group(1)
            if re.search(r'\*\*预期结果\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL) else "")
        
        cases_list.append(TestCaseItem(
            用例编号=case_id,
            用例标题=case_title,
            优先级=priority,
            测试步骤=steps,
            预期结果=expected,
        ))
    
    if not cases_list:
        return {"success": False, "error": "未解析到测试用例"}
    
    # 构建思维导图数据
    mindmap_data = {
        "title": "测试用例",
        "topics": [{
            "title": "📁 测试用例",
            "children": [{
                "title": f"[{c.优先级}] {c.用例标题}",
                "children": [
                    {"title": "📝 测试步骤", "children": [{"title": f"{i}. {s}"} for i, s in enumerate(c.测试步骤, 1)]},
                    {"title": "✅ 预期结果", "children": [{"title": f"{i}. {r}"} for i, r in enumerate(c.预期结果, 1)]},
                ]
            } for c in cases_list]
        }]
    }
    
    # 保存文件
    output_path = Path(output_dir) if output_dir else Path(__file__).parent.parent / "output"
    output_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_path / f"testcases_{timestamp}_mindmap.json"
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(mindmap_data, f, ensure_ascii=False, indent=2)
    
    return {
        "success": True,
        "file_path": str(filepath),
        "total_cases": len(cases_list),
    }


@tool
def export_to_excel(testcase_text: str, output_dir: Optional[str] = None) -> Dict[str, Any]:
    """将测试用例导出为Excel表格
    
    解析测试用例文本，生成Excel格式的测试用例表格。
    
    Args:
        testcase_text: 测试用例文本（Markdown格式）
        output_dir: 输出目录，默认为 text2case/output
        
    Returns:
        导出结果，包含文件路径和统计信息
    """
    import csv
    
    # 解析测试用例
    cases_data = []
    case_pattern = r'#{2,4}\s*(TC-?\d+)\s*([^\n]*)'
    headers = list(re.finditer(case_pattern, testcase_text))
    
    for idx, match in enumerate(headers):
        case_id = match.group(1)
        case_title = match.group(2).strip() or case_id
        
        start = match.end()
        end = headers[idx + 1].start() if idx + 1 < len(headers) else len(testcase_text)
        content = testcase_text[start:end]
        
        # 提取字段
        priority_match = re.search(r'\*\*优先级\*\*[：:]\s*(\S+)', content)
        priority = priority_match.group(1) if priority_match else "P1"
        
        precondition_match = re.search(r'\*\*前置条件\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL)
        precondition = precondition_match.group(1).strip() if precondition_match else ""
        
        steps_match = re.search(r'\*\*测试步骤\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL)
        steps = steps_match.group(1).strip() if steps_match else ""
        
        expected_match = re.search(r'\*\*预期结果\*\*[：:]([^*]+?)(?=\*\*|$)', content, re.DOTALL)
        expected = expected_match.group(1).strip() if expected_match else ""
        
        cases_data.append({
            "用例编号": case_id,
            "用例标题": case_title,
            "优先级": priority,
            "前置条件": precondition,
            "测试步骤": steps,
            "预期结果": expected,
        })
    
    if not cases_data:
        return {"success": False, "error": "未解析到测试用例"}
    
    # 保存文件
    output_path = Path(output_dir) if output_dir else Path(__file__).parent.parent / "output"
    output_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 尝试使用 openpyxl，否则使用 CSV
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        headers = ["用例编号", "用例标题", "优先级", "前置条件", "测试步骤", "预期结果"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, case in enumerate(cases_data, 2):
            ws.cell(row=row, column=1, value=case["用例编号"])
            ws.cell(row=row, column=2, value=case["用例标题"])
            ws.cell(row=row, column=3, value=case["优先级"])
            ws.cell(row=row, column=4, value=case["前置条件"])
            ws.cell(row=row, column=5, value=case["测试步骤"])
            ws.cell(row=row, column=6, value=case["预期结果"])
        
        filepath = output_path / f"testcases_{timestamp}.xlsx"
        wb.save(str(filepath))
        
    except ImportError:
        # 降级为 CSV
        filepath = output_path / f"testcases_{timestamp}.csv"
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=["用例编号", "用例标题", "优先级", "前置条件", "测试步骤", "预期结果"])
            writer.writeheader()
            writer.writerows(cases_data)
    
    return {
        "success": True,
        "file_path": str(filepath),
        "total_cases": len(cases_data),
    }


# ============== 工具集合 ==============

TESTCASE_TOOLS = [
    select_test_methods,
    generate_test_data,
    validate_testcase_format,
]

EXPORT_TOOLS = [
    export_to_xmind,
    export_to_excel,
]

__all__ = [
    "select_test_methods",
    "generate_test_data",
    "validate_testcase_format",
    "export_to_xmind",
    "export_to_excel",
    "TESTCASE_TOOLS",
    "EXPORT_TOOLS",
]
