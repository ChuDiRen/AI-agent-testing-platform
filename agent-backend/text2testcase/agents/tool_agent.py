"""数据处理专家 (Tool Agent) - 非LLM的确定性任务处理

职责:
- 收集和整理数据 (去重、验证)
- 导出为XMind思维导图
- 导出为Excel表格
- 生成统计报告

设计理念:
- 这不是一个基于LLM的Agent，而是一个自定义节点函数
- 数据处理是确定性任务，不需要"理解"和"创造"
- 固定的执行流程更快、更准确、更可控
- 降低成本（数据处理不消耗LLM Token）
"""
import asyncio
import json
import re
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from difflib import SequenceMatcher

from ..models import TestCaseState, TestCaseSuite, TestCaseModule, TestCaseItem


class MarkdownTestCaseParser:
    """Markdown 测试用例解析器 - 从Markdown格式文本中提取测试用例"""
    
    @staticmethod
    def parse(text: str) -> Tuple[Optional[TestCaseSuite], List[str]]:
        """解析Markdown格式的测试用例
        
        支持格式:
        ### TC-001 用例标题
        **优先级**：P0
        **前置条件**：...
        **测试步骤**：...
        **预期结果**：...
        **测试数据**：...
        """
        errors = []
        cases_list = []
        
        # 使用findall找到所有用例标题位置
        # 支持 ## TC-001, ### TC-001, #### TC-001 格式
        case_header_pattern = r'#{2,4}\s*(TC-?\d+)\s*([^\n]*)'
        headers = list(re.finditer(case_header_pattern, text))
        
        if not headers:
            return None, ["未找到测试用例 (格式: ### TC-001 用例标题)"]
        
        # 解析每个用例
        for idx, match in enumerate(headers):
            case_id = match.group(1)
            case_title = match.group(2).strip() or case_id
            
            # 获取用例内容 (从当前标题到下一个标题之间)
            start = match.end()
            end = headers[idx + 1].start() if idx + 1 < len(headers) else len(text)
            content = text[start:end]
            
            # 提取各字段
            priority = MarkdownTestCaseParser._extract_field(content, r'\*\*优先级\*\*[：:]\s*(\S+)')
            precondition = MarkdownTestCaseParser._extract_field(content, r'\*\*前置条件\*\*[：:]([^*]+?)(?=\*\*|$)')
            steps = MarkdownTestCaseParser._extract_list(content, r'\*\*测试步骤\*\*[：:]([^*]+?)(?=\*\*|$)')
            expected = MarkdownTestCaseParser._extract_list(content, r'\*\*预期结果\*\*[：:]([^*]+?)(?=\*\*|$)')
            test_data = MarkdownTestCaseParser._extract_field(content, r'\*\*测试数据\*\*[：:]([^*]+?)(?=\*\*|---|$)')
            
            # 创建用例
            case = TestCaseItem(
                用例编号=case_id,
                用例标题=case_title,
                优先级=priority or "P1",
                前置条件=precondition.strip() if precondition else "",
                测试步骤=steps,
                预期结果=expected,
                测试数据={"raw": test_data.strip()} if test_data else {}
            )
            cases_list.append(case)
        
        if not cases_list:
            return None, ["解析失败: 未提取到任何测试用例"]
        
        # 构建TestCaseSuite
        suite = TestCaseSuite(测试用例=[
            TestCaseModule(功能模块="测试用例", 测试用例列表=cases_list)
        ])
        
        return suite, errors
    
    @staticmethod
    def _extract_field(text: str, pattern: str) -> Optional[str]:
        """提取单个字段"""
        match = re.search(pattern, text, re.DOTALL)
        if match:
            return match.group(1).strip()
        return None
    
    @staticmethod
    def _extract_list(text: str, pattern: str) -> List[str]:
        """提取列表字段"""
        match = re.search(pattern, text, re.DOTALL)
        if not match:
            return []
        
        content = match.group(1)
        # 按数字编号分割
        items = re.findall(r'\d+\.\s*([^\n]+)', content)
        return [item.strip() for item in items if item.strip()]


class PydanticJSONParser:
    """Pydantic JSON 解析器 - 从文本中提取并验证测试用例"""
    
    @staticmethod
    def extract_json_from_text(text: str) -> Optional[str]:
        """从文本中提取JSON内容
        
        支持多种格式:
        - 纯JSON
        - Markdown代码块中的JSON
        - 混合文本中的JSON
        """
        # 尝试直接解析
        text = text.strip()
        if text.startswith('{') or text.startswith('['):
            return text
        
        # 尝试从Markdown代码块中提取
        json_patterns = [
            r'```json\s*([\s\S]*?)\s*```',  # ```json ... ```
            r'```\s*([\s\S]*?)\s*```',       # ``` ... ```
            r'\{[\s\S]*"测试用例"[\s\S]*\}',  # 直接匹配JSON对象
        ]
        
        for pattern in json_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                try:
                    # 验证是否为有效JSON
                    json.loads(match if isinstance(match, str) else match)
                    return match if isinstance(match, str) else match
                except json.JSONDecodeError:
                    continue
        
        return None
    
    @staticmethod
    def parse_test_cases(json_text: str) -> Tuple[Optional[TestCaseSuite], List[str]]:
        """解析测试用例JSON并验证
        
        Returns:
            (TestCaseSuite, errors): 解析结果和错误列表
        """
        errors = []
        
        try:
            data = json.loads(json_text)
        except json.JSONDecodeError as e:
            errors.append(f"JSON解析错误: {str(e)}")
            return None, errors
        
        # 尝试解析为TestCaseSuite
        try:
            if isinstance(data, dict) and "测试用例" in data:
                suite = TestCaseSuite(**data)
                return suite, errors
            elif isinstance(data, list):
                # 如果是列表，包装为套件
                suite = TestCaseSuite(测试用例=[
                    TestCaseModule(**item) if isinstance(item, dict) else item
                    for item in data
                ])
                return suite, errors
        except Exception as e:
            errors.append(f"数据验证错误: {str(e)}")
        
        return None, errors
    
    @staticmethod
    def parse_from_text(text: str) -> Tuple[Optional[TestCaseSuite], List[str]]:
        """从文本中解析测试用例 (支持JSON和Markdown格式)"""
        # 1. 先尝试JSON解析
        json_text = PydanticJSONParser.extract_json_from_text(text)
        if json_text:
            suite, errors = PydanticJSONParser.parse_test_cases(json_text)
            if suite and suite.total_cases > 0:
                return suite, errors
        
        # 2. 如果JSON解析失败，尝试Markdown解析
        suite, errors = MarkdownTestCaseParser.parse(text)
        if suite and suite.total_cases > 0:
            return suite, errors
        
        return None, ["未找到有效的测试用例内容 (支持JSON或Markdown格式)"]


class TestCaseDeduplicator:
    """测试用例去重器 - 基于语义相似度"""
    
    @staticmethod
    def similarity(a: str, b: str) -> float:
        """计算两个字符串的相似度 (0-1)"""
        return SequenceMatcher(None, a, b).ratio()
    
    @staticmethod
    def deduplicate_cases(cases: List[TestCaseItem], threshold: float = 0.85) -> List[TestCaseItem]:
        """去重测试用例
        
        Args:
            cases: 测试用例列表
            threshold: 相似度阈值，超过此值认为是重复
            
        Returns:
            去重后的测试用例列表
        """
        if not cases:
            return []
        
        unique_cases = []
        for case in cases:
            is_duplicate = False
            case_text = f"{case.用例标题} {' '.join(case.测试步骤)}"
            
            for unique_case in unique_cases:
                unique_text = f"{unique_case.用例标题} {' '.join(unique_case.测试步骤)}"
                if TestCaseDeduplicator.similarity(case_text, unique_text) > threshold:
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                unique_cases.append(case)
        
        return unique_cases
    
    @staticmethod
    def deduplicate_suite(suite: TestCaseSuite, threshold: float = 0.85) -> TestCaseSuite:
        """去重整个测试套件"""
        deduplicated_modules = []
        
        for module in suite.测试用例:
            unique_cases = TestCaseDeduplicator.deduplicate_cases(
                module.测试用例列表, 
                threshold
            )
            if unique_cases:
                deduplicated_modules.append(TestCaseModule(
                    功能模块=module.功能模块,
                    测试用例列表=unique_cases
                ))
        
        return TestCaseSuite(测试用例=deduplicated_modules)


class XMindExporter:
    """XMind 思维导图导出器"""
    
    @staticmethod
    async def export(suite: TestCaseSuite, output_path: str, title: str = "测试用例") -> str:
        """导出为XMind文件
        
        Args:
            suite: 测试用例套件
            output_path: 输出目录
            title: 思维导图标题
            
        Returns:
            生成的文件路径
        """
        try:
            import xmind
            from xmind.core.markerref import MarkerId
        except ImportError:
            # 如果没有xmind库，生成JSON格式的思维导图数据
            return await XMindExporter._export_as_json(suite, output_path, title)
        
        # 创建XMind工作簿
        workbook = xmind.Workbook()
        sheet = workbook.getPrimarySheet()
        sheet.setTitle(title)
        
        # 创建根节点
        root = sheet.getRootTopic()
        root.setTitle(title)
        
        # 添加模块和用例
        for module in suite.测试用例:
            module_topic = root.addSubTopic()
            module_topic.setTitle(f"📁 {module.功能模块}")
            
            for case in module.测试用例列表:
                case_topic = module_topic.addSubTopic()
                case_topic.setTitle(f"[{case.优先级}] {case.用例标题}")
                
                # 添加测试步骤
                if case.测试步骤:
                    steps_topic = case_topic.addSubTopic()
                    steps_topic.setTitle("📝 测试步骤")
                    for i, step in enumerate(case.测试步骤, 1):
                        step_topic = steps_topic.addSubTopic()
                        step_topic.setTitle(f"{i}. {step}")
                
                # 添加预期结果
                if case.预期结果:
                    results_topic = case_topic.addSubTopic()
                    results_topic.setTitle("✅ 预期结果")
                    for i, result in enumerate(case.预期结果, 1):
                        result_topic = results_topic.addSubTopic()
                        result_topic.setTitle(f"{i}. {result}")
        
        # 保存文件
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"testcases_{timestamp}.xmind"
        filepath = output_dir / filename
        
        xmind.save(workbook, str(filepath))
        return str(filepath)
    
    @staticmethod
    async def _export_as_json(suite: TestCaseSuite, output_path: str, title: str) -> str:
        """导出为JSON格式的思维导图数据 (备用方案)"""
        mindmap_data = {
            "title": title,
            "topics": []
        }
        
        for module in suite.测试用例:
            module_data = {
                "title": f"📁 {module.功能模块}",
                "children": []
            }
            
            for case in module.测试用例列表:
                case_data = {
                    "title": f"[{case.优先级}] {case.用例标题}",
                    "children": [
                        {
                            "title": "📝 测试步骤",
                            "children": [{"title": f"{i}. {s}"} for i, s in enumerate(case.测试步骤, 1)]
                        },
                        {
                            "title": "✅ 预期结果",
                            "children": [{"title": f"{i}. {r}"} for i, r in enumerate(case.预期结果, 1)]
                        }
                    ]
                }
                module_data["children"].append(case_data)
            
            mindmap_data["topics"].append(module_data)
        
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"testcases_{timestamp}_mindmap.json"
        filepath = output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(mindmap_data, f, ensure_ascii=False, indent=2)
        
        return str(filepath)


class ExcelExporter:
    """Excel 表格导出器"""
    
    @staticmethod
    async def export(suite: TestCaseSuite, output_path: str, title: str = "测试用例") -> str:
        """导出为Excel文件
        
        Args:
            suite: 测试用例套件
            output_path: 输出目录
            title: 工作表标题
            
        Returns:
            生成的文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            # 如果没有openpyxl库，生成CSV格式
            return await ExcelExporter._export_as_csv(suite, output_path, title)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel工作表名最长31字符
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = ["功能模块", "用例编号", "用例标题", "优先级", "前置条件", "测试步骤", "预期结果", "测试数据"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        row = 2
        for module in suite.测试用例:
            for case in module.测试用例列表:
                ws.cell(row=row, column=1, value=module.功能模块).border = thin_border
                ws.cell(row=row, column=2, value=case.用例编号).border = thin_border
                ws.cell(row=row, column=3, value=case.用例标题).border = thin_border
                ws.cell(row=row, column=4, value=case.优先级).border = thin_border
                ws.cell(row=row, column=5, value=case.前置条件).border = thin_border
                ws.cell(row=row, column=6, value="\n".join(case.测试步骤)).border = thin_border
                ws.cell(row=row, column=7, value="\n".join(case.预期结果)).border = thin_border
                ws.cell(row=row, column=8, value=json.dumps(case.测试数据, ensure_ascii=False)).border = thin_border
                
                # 设置对齐方式
                for col in range(1, 9):
                    ws.cell(row=row, column=col).alignment = cell_alignment
                
                row += 1
        
        # 调整列宽
        column_widths = [15, 12, 30, 8, 25, 40, 40, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 保存文件
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"testcases_{timestamp}.xlsx"
        filepath = output_dir / filename
        
        wb.save(str(filepath))
        return str(filepath)
    
    @staticmethod
    async def _export_as_csv(suite: TestCaseSuite, output_path: str, title: str) -> str:
        """导出为CSV格式 (备用方案)"""
        import csv
        
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"testcases_{timestamp}.csv"
        filepath = output_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 写入表头
            writer.writerow(["功能模块", "用例编号", "用例标题", "优先级", "前置条件", "测试步骤", "预期结果", "测试数据"])
            
            # 写入数据
            for module in suite.测试用例:
                for case in module.测试用例列表:
                    writer.writerow([
                        module.功能模块,
                        case.用例编号,
                        case.用例标题,
                        case.优先级,
                        case.前置条件,
                        " | ".join(case.测试步骤),
                        " | ".join(case.预期结果),
                        json.dumps(case.测试数据, ensure_ascii=False)
                    ])
        
        return str(filepath)


class StatisticsGenerator:
    """统计报告生成器"""
    
    @staticmethod
    def generate(suite: TestCaseSuite) -> Dict[str, Any]:
        """生成统计报告
        
        Args:
            suite: 测试用例套件
            
        Returns:
            统计信息字典
        """
        # 基础统计
        total_cases = suite.total_cases
        total_modules = suite.modules_count
        
        # 优先级分布
        priority_dist = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
        
        # 模块统计
        module_stats = []
        
        for module in suite.测试用例:
            module_case_count = len(module.测试用例列表)
            module_priority_dist = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
            
            for case in module.测试用例列表:
                priority = case.优先级
                priority_dist[priority] = priority_dist.get(priority, 0) + 1
                module_priority_dist[priority] = module_priority_dist.get(priority, 0) + 1
            
            module_stats.append({
                "模块名称": module.功能模块,
                "用例数量": module_case_count,
                "优先级分布": module_priority_dist
            })
        
        return {
            "总用例数": total_cases,
            "模块数量": total_modules,
            "优先级分布": priority_dist,
            "模块统计": module_stats,
            "生成时间": datetime.now().isoformat()
        }


async def collect_and_organize_data(state: TestCaseState) -> Dict[str, Any]:
    """收集和整理数据 (步骤1)
    
    Args:
        state: 当前状态
        
    Returns:
        整理后的数据
    """
    # 解析测试用例
    parser = PydanticJSONParser()
    suite, errors = parser.parse_from_text(state.testcases)
    
    if errors:
        print(f"  ⚠️ 解析警告: {errors}")
    
    if suite:
        print(f"  ✅ 解析成功: 共 {suite.total_cases} 个用例")
    else:
        print(f"  ❌ 解析失败，尝试创建空结构")
        # 如果解析失败，尝试创建基础结构
        suite = TestCaseSuite(测试用例=[
            TestCaseModule(
                功能模块="未分类",
                测试用例列表=[]
            )
        ])
    
    # 去重
    deduplicator = TestCaseDeduplicator()
    original_count = suite.total_cases
    suite = deduplicator.deduplicate_suite(suite)
    deduplicated_count = suite.total_cases
    
    return {
        "suite": suite,
        "original_count": original_count,
        "deduplicated_count": deduplicated_count,
        "removed_count": original_count - deduplicated_count,
        "parse_errors": errors
    }


async def export_to_xmind(suite: TestCaseSuite, output_path: str, title: str = "测试用例") -> Dict[str, Any]:
    """导出为XMind (步骤2)"""
    exporter = XMindExporter()
    filepath = await exporter.export(suite, output_path, title)
    return {
        "xmind_path": filepath,
        "success": True
    }


async def export_to_excel(suite: TestCaseSuite, output_path: str, title: str = "测试用例") -> Dict[str, Any]:
    """导出为Excel (步骤3)"""
    exporter = ExcelExporter()
    filepath = await exporter.export(suite, output_path, title)
    return {
        "excel_path": filepath,
        "success": True
    }


async def test_tool_node(state: TestCaseState) -> Dict[str, Any]:
    """数据处理专家节点 - 按固定顺序执行任务
    
    这是一个自定义节点函数，不是基于LLM的Agent。
    
    执行流程:
    1. 收集和整理数据（去重、验证）
    2. 并行导出为XMind思维导图和Excel表格
    3. 生成统计报告
    4. 返回所有下载链接
    
    Args:
        state: 当前状态
        
    Returns:
        更新后的状态字典
    """
    print("\n[5/5] 数据处理专家开始工作...")
    
    # 确定输出目录
    output_dir = Path(__file__).parent.parent / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 步骤1: 收集和整理数据
    print("  📊 步骤1: 收集和整理数据...")
    collect_result = await collect_and_organize_data(state)
    suite = collect_result["suite"]
    
    if collect_result["removed_count"] > 0:
        print(f"  ✅ 去重完成: 移除 {collect_result['removed_count']} 个重复用例")
    
    # 步骤2 & 3: 并行导出XMind和Excel
    print("  📁 步骤2&3: 并行导出文档...")
    
    xmind_task = export_to_xmind(suite, str(output_dir), "测试用例思维导图")
    excel_task = export_to_excel(suite, str(output_dir), "测试用例")
    
    xmind_result, excel_result = await asyncio.gather(
        xmind_task,
        excel_task,
        return_exceptions=True
    )
    
    # 处理结果
    xmind_path = ""
    excel_path = ""
    
    if isinstance(xmind_result, dict) and xmind_result.get("success"):
        xmind_path = xmind_result["xmind_path"]
        print(f"  ✅ XMind导出完成: {xmind_path}")
    elif isinstance(xmind_result, Exception):
        print(f"  ⚠️ XMind导出失败: {xmind_result}")
    
    if isinstance(excel_result, dict) and excel_result.get("success"):
        excel_path = excel_result["excel_path"]
        print(f"  ✅ Excel导出完成: {excel_path}")
    elif isinstance(excel_result, Exception):
        print(f"  ⚠️ Excel导出失败: {excel_result}")
    
    # 步骤4: 生成统计报告
    print("  📈 步骤4: 生成统计报告...")
    statistics = StatisticsGenerator.generate(suite)
    
    print(f"\n  📊 统计信息:")
    print(f"     - 总用例数: {statistics['总用例数']}")
    print(f"     - 模块数量: {statistics['模块数量']}")
    print(f"     - 优先级分布: {statistics['优先级分布']}")
    
    # 构建返回结果
    updates = {
        "xmind_path": xmind_path,
        "excel_path": excel_path,
        "statistics": statistics,
        "current_phase": "process",
        "process_completed": True,
    }
    
    print("\n✅ 数据处理专家完成! 所有处理步骤已完成")
    
    return updates


async def run_tool_agent(state: TestCaseState) -> Dict[str, Any]:
    """运行数据处理专家
    
    这是对外暴露的接口，与其他Agent保持一致的调用方式。
    
    Args:
        state: 当前状态
        
    Returns:
        更新后的状态字典
    """
    return await test_tool_node(state)
