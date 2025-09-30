# Copyright (c) 2025 左岚. All rights reserved.
"""
RBAC用户Service
实现用户相关的RBAC业务逻辑
"""

import io
from typing import List, Optional, Tuple, Dict, Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

from app.core.logger import get_logger
from app.core.security import get_password_hash, verify_password
from app.entity.role import Role
from app.entity.user import User
from app.entity.user_role import UserRole
from app.entity.department import Department
from app.repository.base import BaseRepository
from app.repository.role_repository import RoleRepository
from app.repository.user_role_repository import UserRoleRepository
from app.repository.department_repository import DepartmentRepository

logger = get_logger(__name__)


class RBACUserService:
    """
    RBAC用户Service类
    提供用户相关的RBAC业务逻辑处理
    """

    def __init__(self, db: Session):
        """
        初始化RBAC用户Service

        Args:
            db: 数据库会话
        """
        self.db = db
        self.user_repository = BaseRepository(db, User)
        self.user_role_repository = UserRoleRepository(db)
        self.role_repository = RoleRepository(db)
        self.department_repository = DepartmentRepository(db)

    def create_user(self, username: str, password: str, email: str = None,
                   mobile: str = None, dept_id: int = None, ssex: str = None,
                   avatar: str = None, description: str = None) -> User:
        """
        创建用户

        Args:
            username: 用户名
            password: 明文密码
            email: 邮箱
            mobile: 手机号
            dept_id: 部门ID
            ssex: 性别，'0'男 '1'女 '2'保密
            avatar: 头像
            description: 描述

        Returns:
            创建的用户对象

        Raises:
            ValueError: 用户名已存在
        """
        # 检查用户名是否已存在（只检查未删除的用户）
        existing_user = self.db.query(User).filter(
            User.username == username,
            User.is_deleted == 0
        ).first()
        if existing_user:
            raise ValueError(f"用户名 '{username}' 已存在")

        # 加密密码
        hashed_password = get_password_hash(password)

        # 创建用户
        user = User(
            username=username,
            password=hashed_password,
            email=email,
            mobile=mobile,
            dept_id=dept_id,
            ssex=ssex,
            avatar=avatar,
            description=description
        )

        created_user = self.user_repository.create(user)
        logger.info(f"Created user: {username}")
        return created_user

    def get_user_by_id(self, user_id: int) -> Optional[User]:
        """
        根据ID获取用户

        Args:
            user_id: 用户ID

        Returns:
            用户对象或None
        """
        return self.user_repository.get_by_id(user_id)

    def get_user_by_username(self, username: str) -> Optional[User]:
        """
        根据用户名获取用户

        Args:
            username: 用户名

        Returns:
            用户对象或None
        """
        return self.db.query(User).filter(
            User.username == username,
            User.is_deleted == 0
        ).first()

    async def update_last_login(self, user_id: int):
        """
        更新用户最后登录时间

        Args:
            user_id: 用户ID
        """
        from datetime import datetime
        user = self.get_user_by_id(user_id)
        if user:
            user.last_login_time = datetime.utcnow()
            self.db.commit()
            logger.info(f"Updated last login time for user: {user_id}")

    async def get_user_with_roles(self, user_id: int):
        """
        获取用户及其角色信息

        Args:
            user_id: 用户ID

        Returns:
            用户对象（包含角色信息）
        """
        from app.entity.role import Role
        from app.entity.user_role import UserRole

        user = self.db.query(User).filter(User.id == user_id, User.is_deleted == 0).first()
        if user:
            # 加载角色信息
            user_roles = self.db.query(UserRole).filter(UserRole.user_id == user_id).all()
            user.roles = [self.db.query(Role).filter(Role.id == ur.role_id).first() for ur in user_roles]
        return user

    async def get_user_menus(self, user_id: int):
        """
        获取用户菜单权限

        Args:
            user_id: 用户ID

        Returns:
            菜单列表
        """
        from app.entity.menu import Menu
        from app.entity.role_menu import RoleMenu
        from app.entity.user_role import UserRole

        # 获取用户的所有角色
        user_roles = self.db.query(UserRole).filter(UserRole.user_id == user_id).all()
        role_ids = [ur.role_id for ur in user_roles]

        # 获取角色的所有菜单
        menu_ids = set()
        for role_id in role_ids:
            role_menus = self.db.query(RoleMenu).filter(RoleMenu.role_id == role_id).all()
            menu_ids.update([rm.menu_id for rm in role_menus])

        # 获取菜单详情
        menus = self.db.query(Menu).filter(Menu.id.in_(menu_ids), Menu.is_deleted == 0).all()
        return menus

    async def get_user_apis(self, user_id: int):
        """
        获取用户API权限

        Args:
            user_id: 用户ID

        Returns:
            API列表
        """
        from app.entity.api_endpoint import ApiEndpoint
        from app.entity.user_role import UserRole
        from app.repository.role_api_repository import RoleApiRepository

        # 获取用户的所有角色
        user_roles = self.db.query(UserRole).filter(UserRole.user_id == user_id).all()
        role_ids = [ur.role_id for ur in user_roles]

        if not role_ids:
            return []

        # 使用repository获取角色的所有API
        role_api_repository = RoleApiRepository(self.db)
        api_ids = set()
        for role_id in role_ids:
            role_api_ids = role_api_repository.get_api_ids_by_role_id(role_id)
            api_ids.update(role_api_ids)

        if not api_ids:
            return []

        # 获取API详情
        apis = self.db.query(ApiEndpoint).filter(ApiEndpoint.id.in_(api_ids)).all()
        return apis

    async def update_password(self, user_id: int, new_password_hash: str):
        """
        更新用户密码

        Args:
            user_id: 用户ID
            new_password_hash: 新密码哈希
        """
        user = self.get_user_by_id(user_id)
        if user:
            user.password = new_password_hash
            self.db.commit()
            logger.info(f"Updated password for user: {user_id}")

    def authenticate_user(self, username: str, password: str) -> Optional[User]:
        """
        用户认证

        Args:
            username: 用户名
            password: 明文密码

        Returns:
            认证成功的用户对象或None
        """
        user = self.get_user_by_username(username)
        if not user:
            logger.warning(f"User not found: {username}")
            return None

        if not user.is_active():
            logger.warning(f"User is locked: {username}")
            return None

        if not verify_password(password, user.password):
            logger.warning(f"Invalid password for user: {username}")
            return None

        # 更新最后登录时间
        user.update_last_login()
        # 直接提交数据库会话，因为user对象已经被修改
        self.db.commit()
        self.db.refresh(user)

        logger.info(f"User authenticated successfully: {username}")
        return user

    def update_user(self, user_id: int, username: str = None, email: str = None, mobile: str = None,
                   dept_id: int = None, status: str = None, ssex: str = None, avatar: str = None, description: str = None) -> Optional[User]:
        """
        更新用户信息

        Args:
            user_id: 用户ID
            username: 用户名
            email: 邮箱
            mobile: 手机号
            dept_id: 部门ID
            status: 状态
            ssex: 性别
            avatar: 头像
            description: 描述

        Returns:
            更新后的用户对象或None
        """
        user = self.user_repository.get_by_id(user_id)
        if not user:
            logger.warning(f"User not found with id: {user_id}")
            return None

        # 更新用户信息
        user.update_info(
            username=username,
            email=email,
            mobile=mobile,
            dept_id=dept_id,
            status=status,
            ssex=ssex,
            avatar=avatar,
            description=description
        )

        # 提交更改到数据库
        try:
            self.db.commit()
            self.db.refresh(user)
            logger.info(f"Updated user: {user_id}")
            return user
        except Exception as e:
            self.db.rollback()
            logger.error(f"Error updating user {user_id}: {str(e)}")
            raise

    def change_password(self, user_id: int, old_password: str, new_password: str) -> bool:
        """
        修改密码

        Args:
            user_id: 用户ID
            old_password: 旧密码
            new_password: 新密码

        Returns:
            是否修改成功
        """
        user = self.user_repository.get_by_id(user_id)
        if not user:
            logger.warning(f"User not found with id: {user_id}")
            return False

        # 验证旧密码
        if not verify_password(old_password, user.password):
            logger.warning(f"Invalid old password for user: {user_id}")
            return False

        # 设置新密码
        hashed_new_password = get_password_hash(new_password)
        user.change_password(hashed_new_password)

        self.user_repository.update(user)
        logger.info(f"Password changed for user: {user_id}")
        return True

    def delete_user(self, user_id: int) -> bool:
        """
        删除用户

        Args:
            user_id: 用户ID

        Returns:
            是否删除成功
        """
        try:
            # 检查用户是否存在
            user = self.user_repository.get_by_id(user_id)
            if not user:
                logger.warning(f"User not found with id: {user_id}")
                return False

            # 先删除用户角色关联
            self.user_role_repository.delete_by_user_id(user_id)

            # 删除用户
            self.user_repository.delete(user_id)

            self.db.commit()
            logger.info(f"Deleted user: {user_id}")
            return True
        except Exception as e:
            self.db.rollback()
            logger.error(f"Error deleting user {user_id}: {str(e)}")
            return False

    def reset_password(self, user_id: int, new_password: str) -> bool:
        """
        管理员重置用户密码（无需旧密码）

        Args:
            user_id: 用户ID
            new_password: 新密码

        Returns:
            是否重置成功
        """
        try:
            logger.info(f"Starting password reset for user: {user_id}")
            # 查找用户 - 使用id字段而不是user_id
            user = self.db.query(User).filter(User.id == user_id).first()
            if not user:
                logger.warning(f"User not found with id: {user_id}")
                return False

            logger.info(f"Found user: {user.username}, current password hash length: {len(user.password) if user.password else 'None'}")

            # 直接设置新密码（管理员权限，无需验证旧密码）
            hashed_new_password = get_password_hash(new_password)
            logger.info(f"Generated new password hash length: {len(hashed_new_password)}")

            user.change_password(hashed_new_password)
            logger.info(f"Password updated in entity for user: {user_id}")

            # 提交数据库更改
            self.db.commit()
            self.db.refresh(user)
            logger.info(f"Password reset completed successfully for user: {user_id}")
            return True
        except Exception as e:
            logger.error(f"Error in reset_password for user {user_id}: {str(e)}")
            return False

    def lock_user(self, user_id: int) -> bool:
        """
        锁定用户

        Args:
            user_id: 用户ID

        Returns:
            是否锁定成功
        """
        user = self.user_repository.get_by_id(user_id)
        if not user:
            logger.warning(f"User not found with id: {user_id}")
            return False

        user.lock_user()
        self.user_repository.update(user)

        logger.info(f"User locked: {user_id}")
        return True

    def unlock_user(self, user_id: int) -> bool:
        """
        解锁用户

        Args:
            user_id: 用户ID

        Returns:
            是否解锁成功
        """
        user = self.user_repository.get_by_id(user_id)
        if not user:
            logger.warning(f"User not found with id: {user_id}")
            return False

        user.unlock_user()
        self.user_repository.update(user)

        logger.info(f"User unlocked: {user_id}")
        return True

    def assign_roles_to_user(self, user_id: int, role_ids: List[int]) -> bool:
        """
        为用户分配角色

        Args:
            user_id: 用户ID
            role_ids: 角色ID列表

        Returns:
            是否分配成功
        """
        try:
            # 检查用户是否存在
            user = self.user_repository.get_by_id(user_id)
            if not user:
                logger.warning(f"User not found with id: {user_id}")
                return False

            # 检查所有角色是否存在
            for role_id in role_ids:
                role = self.role_repository.get_by_id(role_id)
                if not role:
                    logger.warning(f"Role not found with id: {role_id}")
                    return False

            # 分配角色
            self.user_role_repository.assign_roles_to_user(user_id, role_ids)
            self.db.commit()

            logger.info(f"Assigned {len(role_ids)} roles to user: {user_id}")
            return True

        except Exception as e:
            self.db.rollback()
            logger.error(f"Error assigning roles to user {user_id}: {str(e)}")
            return False

    def get_user_roles(self, user_id: int) -> List[Role]:
        """获取单个用户的角色列表  # 注释"""
        return self.user_role_repository.get_roles_by_user_id(user_id)

    def get_roles_for_users(self, user_ids: List[int]) -> dict:
        """批量获取多个用户的角色，返回 {user_id: List[Role]}  # 注释"""
        return self.user_role_repository.get_roles_by_user_ids(user_ids)

    def get_user_permissions(self, user_id: int) -> List[str]:
        """
        获取用户的权限标识列表

        Args:
            user_id: 用户ID

        Returns:
            权限标识列表
        """
        from app.service.menu_service import MenuService

        menu_service = MenuService(self.db)
        return menu_service.get_user_permissions(user_id)

    def has_permission(self, user_id: int, permission: str) -> bool:
        """
        检查用户是否有指定权限

        Args:
            user_id: 用户ID
            permission: 权限标识

        Returns:
            是否有权限
        """
        user_permissions = self.get_user_permissions(user_id)
        return permission in user_permissions

    def get_all_users(self) -> List[User]:
        """获取所有用户（不带筛选，谨慎使用）  # 注释"""
        return self.user_repository.get_all()

    def query_users(self, page: int, size: int,
                    username: Optional[str] = None,
                    dept_id: Optional[int] = None,
                    status: Optional[str] = None,
                    ssex: Optional[str] = None) -> Tuple[List[User], int]:
        """基于数据库的分页与筛选，预加载部门与角色  # 注释"""
        # 统一条件
        conditions = [User.is_deleted == 0]
        if username:
            like = f"%{username}%"
            conditions.append(or_(User.username.like(like), User.email.like(like), User.mobile.like(like)))
        if dept_id is not None:
            conditions.append(User.dept_id == dept_id)
        if status is not None:
            conditions.append(User.status == status)
        if ssex is not None:
            conditions.append(User.ssex == ssex)

        # 统计总数
        total = self.db.query(func.count(User.id)).filter(*conditions).scalar() or 0

        # 预加载：部门 + 用户角色->角色
        q = self.db.query(User).options(
            selectinload(User.department),
            selectinload(User.user_roles).selectinload(UserRole.role)
        ).filter(*conditions).order_by(User.id)

        start = (page - 1) * size
        users = q.offset(start).limit(size).all()
        return users, total

    def search_users(self, keyword: str) -> List[User]:
        """
        搜索用户

        Args:
            keyword: 搜索关键词

        Returns:
            匹配的用户列表
        """
        # 这里可以根据用户名、邮箱等字段搜索
        # 简化实现，只按用户名搜索
        users = self.user_repository.get_all()
        return [user for user in users if keyword.lower() in user.username.lower()]

    def get_role_menus(self, role_id: int):
        """
        获取角色的菜单权限

        Args:
            role_id: 角色ID

        Returns:
            角色菜单关联列表
        """
        from app.repository.role_menu_repository import RoleMenuRepository
        role_menu_repository = RoleMenuRepository(self.db)
        return role_menu_repository.get_menus_by_role_id(role_id)

    def get_menus_by_ids(self, menu_ids: List[int]):
        """
        根据菜单ID列表获取菜单

        Args:
            menu_ids: 菜单ID列表

        Returns:
            菜单列表
        """
        from app.entity.menu import Menu
        return self.db.query(Menu).filter(Menu.id.in_(menu_ids)).all()  # 修复：使用正确的属性名

    def get_all_menus(self):
        """
        获取所有菜单

        Returns:
            菜单列表
        """
        from app.entity.menu import Menu
        return self.db.query(Menu).all()

    def get_users_by_role(self, role_id: int):
        """
        获取拥有指定角色的用户

        Args:
            role_id: 角色ID

        Returns:
            用户角色关联列表
        """
        return self.user_role_repository.get_users_by_role_id(role_id)

    def assign_role_to_user(self, user_id: int, role_id: int) -> bool:
        """
        为用户分配单个角色

        Args:
            user_id: 用户ID
            role_id: 角色ID

        Returns:
            是否分配成功
        """
        try:
            self.user_role_repository.assign_role_to_user(user_id, role_id)
            self.db.commit()
            logger.info(f"Assigned role {role_id} to user {user_id}")
            return True
        except Exception as e:
            self.db.rollback()
            logger.error(f"Error assigning role to user: {str(e)}")
            return False

    def clear_user_roles(self, user_id: int) -> bool:
        """
        清除用户的所有角色

        Args:
            user_id: 用户ID

        Returns:
            是否清除成功
        """
        try:
            self.user_role_repository.remove_all_roles_from_user(user_id)
            self.db.commit()
            logger.info(f"Cleared all roles for user {user_id}")
            return True
        except Exception as e:
            self.db.rollback()
            logger.error(f"Error clearing user roles: {str(e)}")
            return False

    def export_users_to_excel(self, dept_id: Optional[int] = None, status: Optional[str] = None,
                              ssex: Optional[str] = None, include_roles: bool = True,
                              user_ids: Optional[List[int]] = None) -> bytes:
        """
        导出用户数据到Excel文件

        Args:
            dept_id: 部门ID筛选
            status: 状态筛选
            ssex: 性别筛选
            include_roles: 是否包含角色信息
            user_ids: 指定用户ID列表（如果提供，则只导出这些用户）

        Returns:
            Excel文件的字节数据
        """
        try:
            # 构建查询条件
            query = self.db.query(User).filter(User.is_deleted == 0)  # 排除已删除用户

            # 如果指定了用户ID列表，则只导出这些用户
            if user_ids:
                query = query.filter(User.id.in_(user_ids))
            else:
                # 否则根据筛选条件导出
                if dept_id:
                    query = query.filter(User.dept_id == dept_id)
                if status is not None:
                    query = query.filter(User.status == status)
                if ssex is not None:
                    query = query.filter(User.ssex == ssex)

            users = query.all()
            
            # 获取部门信息映射
            departments = {dept.id: dept.dept_name for dept in self.department_repository.get_all(limit=1000)}
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "用户列表"
            
            # 定义表头
            headers = [
                "用户ID", "用户名", "邮箱", "手机号", "部门", "性别", "状态", 
                "头像", "描述", "创建时间", "最后登录时间"
            ]
            
            if include_roles:
                headers.append("角色")
                
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
            # 填充数据
            for row_num, user in enumerate(users, 2):
                # 性别映射
                sex_map = {"0": "男", "1": "女", "2": "保密"}
                status_map = {"0": "禁用", "1": "启用"}
                
                row_data = [
                    user.user_id,
                    user.username,
                    user.email or "",
                    user.mobile or "",
                    departments.get(user.dept_id, ""),
                    sex_map.get(user.ssex, ""),
                    status_map.get(user.status, ""),
                    user.avatar or "",
                    user.description or "",
                    user.create_time.strftime("%Y-%m-%d %H:%M:%S") if user.create_time else "",
                    user.last_login_time.strftime("%Y-%m-%d %H:%M:%S") if user.last_login_time else ""
                ]
                
                if include_roles:
                    # 获取用户角色
                    roles = self.get_user_roles(user.user_id)
                    role_names = ", ".join([role.role_name for role in roles])
                    row_data.append(role_names)
                
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                    
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Exported {len(users)} users to Excel")
            return output.read()
            
        except Exception as e:
            logger.error(f"Error exporting users to Excel: {str(e)}")
            raise
            
    def import_users_from_excel(self, file_content: bytes, update_existing: bool = False) -> Dict[str, Any]:
        """
        从Excel文件导入用户数据
        
        Args:
            file_content: Excel文件内容
            update_existing: 是否更新已存在的用户
            
        Returns:
            导入结果统计
        """
        try:
            # 加载Excel文件
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active
            
            total_count = 0
            success_count = 0
            failed_count = 0
            error_messages = []
            
            # 获取部门映射（部门名称 -> 部门ID）
            departments = {dept.dept_name: dept.id for dept in self.department_repository.get_all(limit=1000)}
            
            # 读取数据（跳过表头）
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # 跳过空行
                    continue
                    
                total_count += 1
                
                try:
                    # 解析行数据
                    username = str(row[1]).strip() if row[1] else ""
                    email = str(row[2]).strip() if row[2] else None
                    mobile = str(row[3]).strip() if row[3] else None
                    dept_name = str(row[4]).strip() if row[4] else ""
                    ssex_text = str(row[5]).strip() if row[5] else "保密"
                    status_text = str(row[6]).strip() if row[6] else "启用"
                    description = str(row[8]).strip() if row[8] else None
                    
                    # 验证必填字段
                    if not username:
                        error_messages.append(f"第{row_num}行：用户名不能为空")
                        failed_count += 1
                        continue
                        
                    # 转换性别
                    sex_map = {"男": "0", "女": "1", "保密": "2"}
                    ssex = sex_map.get(ssex_text, "2")
                    
                    # 转换状态
                    status_map = {"启用": "1", "禁用": "0"}
                    status = status_map.get(status_text, "1")
                    
                    # 获取部门ID
                    dept_id = departments.get(dept_name) if dept_name else None
                    
                    # 检查用户是否已存在（只检查未删除的用户）
                    existing_user = self.db.query(User).filter(
                        User.username == username,
                        User.is_deleted == 0
                    ).first()

                    # 检查是否有已删除的同名用户
                    deleted_user = self.db.query(User).filter(
                        User.username == username,
                        User.is_deleted == 1
                    ).first()

                    if existing_user:
                        if update_existing:
                            # 更新现有用户
                            existing_user.email = email
                            existing_user.mobile = mobile
                            existing_user.dept_id = dept_id
                            existing_user.ssex = ssex
                            existing_user.status = status
                            existing_user.description = description
                            self.db.commit()
                            success_count += 1
                        else:
                            error_messages.append(f"第{row_num}行：用户名 '{username}' 已存在")
                            failed_count += 1
                    elif deleted_user:
                        # 恢复已删除的用户
                        deleted_user.email = email
                        deleted_user.mobile = mobile
                        deleted_user.dept_id = dept_id
                        deleted_user.ssex = ssex
                        deleted_user.status = status
                        deleted_user.description = description
                        deleted_user.password = get_password_hash("123456")  # 重置密码
                        deleted_user.restore()  # 恢复用户（设置is_deleted=0）
                        self.db.commit()
                        success_count += 1
                    else:
                        # 创建新用户
                        new_user = User(
                            username=username,
                            password=get_password_hash("123456"),  # 默认密码
                            email=email,
                            mobile=mobile,
                            dept_id=dept_id,
                            ssex=ssex,
                            avatar=None,
                            description=description
                        )
                        new_user.status = status  # 设置状态
                        self.db.add(new_user)
                        self.db.commit()
                        success_count += 1
                        
                except Exception as e:
                    error_messages.append(f"第{row_num}行：{str(e)}")
                    failed_count += 1
                    self.db.rollback()
                    
            result = {
                "total_count": total_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "error_messages": error_messages
            }
            
            logger.info(f"Import completed: {success_count}/{total_count} users imported successfully")
            return result
            
        except Exception as e:
            logger.error(f"Error importing users from Excel: {str(e)}")
            raise

    async def get_department_by_id(self, dept_id: int) -> Optional[Department]:
        """
        根据ID获取部门信息

        Args:
            dept_id: 部门ID

        Returns:
            部门对象或None
        """
        try:
            return self.department_repository.get_by_id(dept_id)
        except Exception as e:
            logger.error(f"Error getting department by id {dept_id}: {str(e)}")
            return None
